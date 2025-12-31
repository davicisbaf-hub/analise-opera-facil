# main.py (ou o nome do seu arquivo principal)

from openpyxl import load_workbook
from dotenv import load_dotenv
import pandas as pd
import os
import re
import sys
from pathlib import Path

load_dotenv()

# ============================================================================
# FUNÇÕES PARA PROCESSAR RELATÓRIOS
# ============================================================================

def criar_planilha_municipio_colunas(caminho_arquivo):
    """
    Cria uma planilha onde cada município tem 3 colunas:
    1. Paciente [municipio]
    2. [municipio]  (apenas o nome do município)
    3. Quantidade [municipio]
    """
    
    wb = load_workbook(caminho_arquivo)
    ws = wb.active
    
    # Verificar se merged_cells existe e não é None
    merged_cells_dict = {}
    
    # Verificar se a planilha tem células mescladas
    if ws.merged_cells and ws.merged_cells.ranges is not None:
        merged_ranges = ws.merged_cells.ranges
        
        for merged_range in merged_ranges:
            min_col, min_row, max_col, max_row = merged_range.bounds
            valor = ws.cell(row=min_row, column=min_col).value
            
            for row in range(min_row, max_row + 1):
                for col in range(min_col, max_col + 1):
                    merged_cells_dict[(row, col)] = valor
    
    # Restante do código permanece igual
    dados_por_municipio = {}
    
    municipio_atual = None
    estado_atual = None
    
    for row in range(1, ws.max_row + 1):
        
        cell_value = str(ws.cell(row=row, column=1).value) if ws.cell(row=row, column=1).value else ""
        
        if 'MUNICIPIO:' in cell_value:
            if '-' in cell_value:
                partes = cell_value.split('-')
                estado_atual = partes[0].replace('MUNICIPIO:', '').strip()
                municipio_atual = f"{estado_atual} - {partes[1].strip()}"
            continue
        
        
        data_hora = None
        if (row, 1) in merged_cells_dict:
            data_hora = merged_cells_dict[(row, 1)]
        else:
            data_hora = ws.cell(row=row, column=1).value
        
        if data_hora and isinstance(data_hora, str) and re.match(r'\d{2}/\d{2}/\d{4}', data_hora[:10]):
            
            linha_dados = []
            for col in range(1, 11):
                if (row, col) in merged_cells_dict:
                    valor = merged_cells_dict[(row, col)]
                else:
                    valor = ws.cell(row=row, column=col).value
                linha_dados.append(valor)
            
            if len(linha_dados) >= 10 and municipio_atual:
                paciente = linha_dados[1]
                procedimento = linha_dados[3]
                quantidade = linha_dados[4]
                
                if paciente and procedimento:
                    try:
                        qtd = int(float(quantidade)) if quantidade else 1
                    except:
                        qtd = 1
                    
                    
                    if municipio_atual not in dados_por_municipio:
                        dados_por_municipio[municipio_atual] = []
                    
                    dados_por_municipio[municipio_atual].append({
                        'paciente': paciente,
                        'procedimento': procedimento,
                        'quantidade': qtd
                    })
    
    
    if not dados_por_municipio:
        print("AVISO: Nenhum dado encontrado para organizar por municipio")
        return None, []
    
    
    municipios = list(dados_por_municipio.keys())
    
    
    max_registros = max(len(dados) for dados in dados_por_municipio.values())
    
    
    df_data = {}
    
    
    for municipio in municipios:
        dados = dados_por_municipio[municipio]
        
        
        pacientes = []
        procedimentos = []
        quantidades = []
        
        
        for dado in dados:
            pacientes.append(dado['paciente'])
            procedimentos.append(dado['procedimento'])
            quantidades.append(dado['quantidade'])
        
        
        while len(pacientes) < max_registros:
            pacientes.append(None)
            procedimentos.append(None)
            quantidades.append(None)
        
        
        
        df_data[f'Paciente {municipio}'] = pacientes
        df_data[f'{municipio}'] = procedimentos  
        df_data[f'Quantidade {municipio}'] = quantidades
    
    
    df_final = pd.DataFrame(df_data)
    
    
    df_final = df_final.dropna(how='all')
    
    
    colunas_ordenadas = []
    for municipio in municipios:
        colunas_ordenadas.append(f'Paciente {municipio}')
        colunas_ordenadas.append(f'{municipio}')  
        colunas_ordenadas.append(f'Quantidade {municipio}')
    
    df_final = df_final[colunas_ordenadas]
    
    return df_final, municipios


def criar_planilha_dados_detalhados(caminho_arquivo):
    """
    Cria uma planilha com todos os dados detalhados, incluindo município.
    """
    
    wb = load_workbook(caminho_arquivo)
    ws = wb.active
    
    # Verificar se merged_cells existe e não é None
    merged_cells_dict = {}
    
    if ws.merged_cells and ws.merged_cells.ranges is not None:
        merged_ranges = ws.merged_cells.ranges
        
        for merged_range in merged_ranges:
            min_col, min_row, max_col, max_row = merged_range.bounds
            valor = ws.cell(row=min_row, column=min_col).value
            
            for row in range(min_row, max_row + 1):
                for col in range(min_col, max_col + 1):
                    merged_cells_dict[(row, col)] = valor
    
    # Restante do código permanece igual
    dados_detalhados = []
    municipio_atual = None
    estado_atual = None
    
    for row in range(1, ws.max_row + 1):
        
        cell_value = str(ws.cell(row=row, column=1).value) if ws.cell(row=row, column=1).value else ""
        
        if 'MUNICIPIO:' in cell_value:
            if '-' in cell_value:
                partes = cell_value.split('-')
                estado_atual = partes[0].replace('MUNICIPIO:', '').strip()
                municipio_atual = f"{estado_atual} - {partes[1].strip()}"
            continue
        
        
        data_hora = None
        if (row, 1) in merged_cells_dict:
            data_hora = merged_cells_dict[(row, 1)]
        else:
            data_hora = ws.cell(row=row, column=1).value
        
        if data_hora and isinstance(data_hora, str) and re.match(r'\d{2}/\d{2}/\d{4}', data_hora[:10]):
            
            linha_dados = []
            for col in range(1, 11):
                if (row, col) in merged_cells_dict:
                    valor = merged_cells_dict[(row, col)]
                else:
                    valor = ws.cell(row=row, column=col).value
                linha_dados.append(valor)
            
            if len(linha_dados) >= 10 and municipio_atual:
                
                linha_completa = [municipio_atual] + linha_dados
                dados_detalhados.append(linha_completa)
    
    
    if dados_detalhados:
        colunas = ['Municipio', 'Data/Hora', 'Paciente', 'Data Nascimento', 
                  'Procedimento', 'Quantidade', 'Valor Regional', 'Contraste', 
                  'Sedacao', 'Valor SUS', 'Valor Total']
        
        
        if len(dados_detalhados[0]) < len(colunas):
            colunas = colunas[:len(dados_detalhados[0])]
        
        df_detalhado = pd.DataFrame(dados_detalhados, columns=colunas)
        return df_detalhado
    else:
        return None


def processar_relatorio_simplificado(caminho_arquivo):
    """
    Processa o relatório e salva apenas as 2 planilhas solicitadas:
    1. Por Município Colunas (com formato modificado)
    2. Dados Detalhados
    """
    
    print(f"Processando: {caminho_arquivo}")
    
    
    if not os.path.exists(caminho_arquivo):
        print(f"ERRO: Arquivo nao encontrado: {caminho_arquivo}")
        return None
    
    try:
        df_municipio_colunas, municipios_encontrados = criar_planilha_municipio_colunas(caminho_arquivo)
        df_dados_detalhados = criar_planilha_dados_detalhados(caminho_arquivo)
    except Exception as e:
        print(f"ERRO ao processar arquivo {caminho_arquivo}: {str(e)}")
        return None
    
    
    if df_municipio_colunas is None and df_dados_detalhados is None:
        print(f"ERRO: Nenhum dado encontrado no arquivo: {caminho_arquivo}")
        return None
    
    
    output_dir = "../relatorios_simplificados"
    if not os.path.exists(output_dir):
        os.makedirs(output_dir)
    
    
    nome_base = os.path.splitext(os.path.basename(caminho_arquivo))[0]
    nome_arquivo = f'{output_dir}/{nome_base}_SIMPLIFICADO.xlsx'
    
    
    try:
        with pd.ExcelWriter(nome_arquivo, engine='openpyxl') as writer:
            
            if df_municipio_colunas is not None:
                df_municipio_colunas.to_excel(writer, sheet_name='Por Municipio Colunas', index=False)
                print("   OK: Planilha 'Por Municipio Colunas' criada (formato modificado)")
            
            
            if df_dados_detalhados is not None:
                df_dados_detalhados.to_excel(writer, sheet_name='Dados Detalhados', index=False)
                print("   OK: Planilha 'Dados Detalhados' criada")
        
        print(f"OK: Arquivo salvo: {nome_arquivo}")
        
    except Exception as e:
        print(f"ERRO ao salvar arquivo: {str(e)}")
        return None
    
    
    print("\nRESUMO DO ARQUIVO:")
    print("-" * 60)
    
    if df_municipio_colunas is not None:
        print(f"• Planilha 'Por Municipio Colunas':")
        print(f"  - Municipios: {len(municipios_encontrados)}")
        print(f"  - Colunas: {len(df_municipio_colunas.columns)}")
        print(f"  - Linhas: {len(df_municipio_colunas)}")
    
    if df_dados_detalhados is not None:
        print(f"• Planilha 'Dados Detalhados':")
        print(f"  - Registros: {len(df_dados_detalhados)}")
        print(f"  - Colunas: {len(df_dados_detalhados.columns)}")
    
    
    if df_municipio_colunas is not None and len(municipios_encontrados) > 0:
        print(f"\nEXEMPLO DO NOVO FORMATO 'Por Municipio Colunas':")
        print("=" * 100)
        
        
        print(f"\nEstrutura das colunas (primeiros 2 municipios como exemplo):")
        for municipio in municipios_encontrados[:2]:
            print(f"  • Paciente {municipio}")
            print(f"  • {municipio}")  
            print(f"  • Quantidade {municipio}")
        
        
        if len(df_municipio_colunas) > 0:
            print(f"\nPrimeiras 2 linhas de exemplo:")
            print("-" * 80)
            
            for i in range(min(2, len(df_municipio_colunas))):
                print(f"Linha {i+1}:")
                
                
                for municipio in municipios_encontrados[:2]:  
                    paciente_col = f'Paciente {municipio}'
                    
                    if paciente_col in df_municipio_colunas.columns:
                        paciente = df_municipio_colunas.iloc[i][paciente_col]
                        if pd.notna(paciente):
                            print(f"  Municipio: {municipio}")
                            print(f"    Paciente: {paciente}")
                            print(f"    Procedimento: {df_municipio_colunas.iloc[i][f'{municipio}']}")
                            print(f"    Quantidade: {df_municipio_colunas.iloc[i][f'Quantidade {municipio}']}")
                print("-" * 40)
    
    return {
        'df_municipio_colunas': df_municipio_colunas,
        'df_dados_detalhados': df_dados_detalhados,
        'municipios_encontrados': municipios_encontrados,
        'arquivo_saida': nome_arquivo
    }


def processar_todos_arquivos_simplificado():
    """
    Processa todos os arquivos listados na variável de ambiente
    mantendo apenas as 2 planilhas solicitadas.
    """
    
    arquivos_str = os.getenv("separarArquivo", "")
    
    if not arquivos_str:
        arquivos = [
            "../separarRelatorio/separarNeomater.xlsx",
            "../separarRelatorio/separarNeotin.xlsx", 
            "../separarRelatorio/separarPediatrico.xlsx"
        ]
    else:
        
        arquivos_str = arquivos_str.strip()
        if arquivos_str.startswith('["') and arquivos_str.endswith('"]'):
            arquivos_str = arquivos_str[2:-2]
        
        arquivos = [arquivo.strip() for arquivo in arquivos_str.split(',') if arquivo.strip()]
    
    print(f"Total de arquivos a processar: {len(arquivos)}")
    
    resultados = []
    arquivos_processados = 0
    
    for arquivo in arquivos:
        print("\n" + "=" * 80)
        print(f"PROCESSANDO: {arquivo}")
        print("=" * 80)
        
        if not os.path.exists(arquivo):
            print(f"ERRO: Arquivo nao encontrado: {arquivo}")
            continue
        
        try:
            resultado = processar_relatorio_simplificado(arquivo)
            if resultado is not None:
                resultados.append(resultado)
                arquivos_processados += 1
                print(f"OK: Concluido: {arquivo}")
            
        except Exception as e:
            print(f"ERRO ao processar {arquivo}: {str(e)}")
            import traceback
            traceback.print_exc()
    
    print("\n" + "=" * 80)
    print("PROCESSAMENTO CONCLUIDO!")
    print("=" * 80)
    
    
    if resultados:
        print(f"\nRESUMO GERAL:")
        print("-" * 60)
        
        total_municipios = 0
        total_registros = 0
        
        for resultado in resultados:
            arquivo = resultado['arquivo_saida']
            num_municipios = len(resultado['municipios_encontrados'])
            
            num_registros_detalhados = 0
            if resultado['df_dados_detalhados'] is not None:
                num_registros_detalhados = len(resultado['df_dados_detalhados'])
            
            print(f"Arquivo: {os.path.basename(arquivo)}:")
            print(f"   • Municipios encontrados: {num_municipios}")
            print(f"   • Registros detalhados: {num_registros_detalhados}")
            
            total_municipios += num_municipios
            total_registros += num_registros_detalhados
        
        print(f"\nTOTAIS:")
        print(f"   • Arquivos processados: {arquivos_processados}/{len(arquivos)}")
        print(f"   • Total de municipios distintos: {total_municipios}")
        print(f"   • Total de registros detalhados: {total_registros}")
        
        
        if resultados:
            print(f"\nESTRUTURA DOS ARQUIVOS GERADOS:")
            print("-" * 60)
            print(f"Todos os arquivos foram salvos na pasta: relatorios_simplificados/")
            print(f"Cada arquivo contem 2 planilhas:")
            print(f"  1. 'Por Municipio Colunas' - Formato com colunas por municipio")
            print(f"      • Paciente [municipio]")
            print(f"      • [municipio] (apenas o nome)")
            print(f"      • Quantidade [municipio]")
            print(f"  2. 'Dados Detalhados' - Dados completos organizados")
    
    return resultados

# ============================================================================
# FUNÇÕES PARA ANALISAR PROCEDIMENTOS (INCLUÍDAS AQUI PARA EVITAR IMPORTAÇÃO)
# ============================================================================

def pacote_otorrino():
    """Retorna lista de procedimentos otorrinolaringológicos"""
    return [
        "4.09.04.030-1", "3.02.04.020-1", "3.01.02.002-1",
        "3.01.01.023-1", "3.01.01.022-2", "3.01.01.004-1"
    ]

def pacote_geral():
    """Retorna lista de procedimentos de cirurgia geral"""
    return [
        "4.09.04.030-1", "3.02.04.020-1", "3.01.02.002-1",
        "3.01.01.023-1", "3.01.01.022-2", "3.01.01.004-1"
    ]

def pacote_oftalmo():
    """Retorna lista de procedimentos oftalmológicos"""
    return [
        "4.09.04.030-1", "3.02.04.020-1", "3.01.02.002-1",
        "3.01.01.023-1", "3.01.01.022-2", "3.01.01.004-1"
    ]

def pacote_hispospadia():
    """Retorna procedimento de hispospadia"""
    return ["3.04.01.012-3"]

def pacote_inguinal():
    """Retorna procedimento de hernioplastia inguinal"""
    return ["3.04.01.013-1"]

def pacote_hidrocele():
    """Retorna procedimento de hidrocele"""
    return ["3.04.01.014-0"]

def pacote_adeno():
    """Retorna procedimento de adenoidectomia"""
    return ["3.03.01.003-9"]

def pacote_amig():
    """Retorna procedimento de amigdalectomia"""
    return ["3.03.01.001-2"]

def pacote_amig_adeno():
    """Retorna procedimento de amigdalectomia com adenoidectomia"""
    return ["3.03.01.004-7"]

def pacote_estrabismo():
    """Retorna procedimento de estrabismo"""
    return ["3.06.02.001-3"]

def pacote_nasal():
    """Retorna procedimento de septo nasal"""
    return ["3.03.04.011-6"]

def pacote_orqui():
    """Retorna procedimento de orquidopexia"""
    return ["3.04.01.002-5"]

def pacote_plastica():
    """Retorna procedimento de plástica de pênis"""
    return ["3.04.01.016-6"]

def pacote_postec():
    """Retorna procedimento de postectomia"""
    return ["3.04.01.015-8"]

def pacote_umbilical():
    """Retorna procedimento de hernioplastia umbilical"""
    return ["3.04.01.009-3"]

# ============================================================================
# FUNÇÃO ANALISAR_NEOMATER
# ============================================================================

def analisar_neomater():
    # Determinar o diretório de execução atual
    if getattr(sys, 'frozen', False):
        # Executando a partir de um executável (.exe)
        base_dir = Path(sys.executable).parent
    else:
        # Executando a partir do script Python
        base_dir = Path(__file__).parent
    
    arquivo = base_dir / "../relatorios_simplificados/separarNeomater_SIMPLIFICADO.xlsx"
    
    # Criar diretórios de saída se não existirem
    output_dir = base_dir / "../Prestador/neomater/resultado"
    output_dir.mkdir(parents=True, exist_ok=True)
    
    municipios = ["RJ - Belford Roxo", "RJ - Duque de Caxias", "RJ - Itaguaí", "RJ - Japeri", 
                  "RJ - Magé", "RJ - Mesquita", "RJ - Nilópolis", "RJ - Nova Iguaçu", 
                  "RJ - Paracambi", "RJ - Queimados", "RJ - Seropédica", "RJ - São João de Meriti"]
    
    # Dicionário para acumular os resultados de todos os municípios
    resultados_por_municipio = {}
    
    # Lista para acumular DataFrames de procedimentos não listados
    nao_listados_dfs = []

    for municipio in municipios:
        try:
            print(f"Processando {municipio}...")
            
            # Verificar se o arquivo existe
            if not arquivo.exists():
                print(f"Arquivo {arquivo} não encontrado!")
                continue
                
            # Inicializar todas as listas de procedimentos
            otorrino = pacote_otorrino()
            geral = pacote_geral()
            oftalmo = pacote_oftalmo()
            hispospadia = pacote_hispospadia()
            inguinal = pacote_inguinal()
            hidrocele = pacote_hidrocele()
            adeno = pacote_adeno()
            amig = pacote_amig()
            amig_adeno = pacote_amig_adeno()
            estrabismo = pacote_estrabismo()
            nasal = pacote_nasal()
            orqui = pacote_orqui()
            plastica = pacote_plastica()
            postec = pacote_postec()
            umbilical = pacote_umbilical()

            tabela = pd.read_excel(arquivo)

            coluna_procedimento = municipio
            coluna_quantidade = f"Quantidade {municipio}"

            if coluna_procedimento not in tabela.columns:
                print(f"Coluna '{coluna_procedimento}' não encontrada em {municipio}!")
                continue
            
            if coluna_quantidade not in tabela.columns:
                print(f"Coluna '{coluna_quantidade}' não encontrada em {municipio}!")
                continue
            
            resultados_municipio = {}
            
            def process_group(procedimentos):
                total = 0
                for procedimento in procedimentos:
                    mask = tabela[coluna_procedimento].astype(str) == str(procedimento)
                    quantidade = tabela.loc[mask, coluna_quantidade].sum()
                    
                    try:
                        quantidade_num = float(quantidade) if not pd.isna(quantidade) else 0
                    except (ValueError, TypeError):
                        quantidade_num = 0

                    if quantidade_num > 0:
                        resultados_municipio[str(procedimento)] = quantidade_num
                        total += quantidade_num
                return total
            
            # Grupos de procedimentos
            grupos = {
                "PACOTE PRÉ-OPERATÓRIO PEDIÁTRICO OTORRINO": otorrino,
                "PACOTE PRÉ-OPERATÓRIO PEDIÁTRICO CIRURGIA GERAL": geral,
                "PACOTE PRÉ-OPERATÓRIO PEDIÁTRICO OFTALMOLOGISTA": oftalmo,
                "ADENOIDECTOMIA PEDIÁTRICO": adeno,
                "AMIGDALECTOMIA - PEDIATRICO": amig,
                "AMIGDALECTOMIA COM ADENOIDECTOMIA - PEDIATRICO": amig_adeno,
                "TRATAMENTO CIRÚRGICO DE PERFURAÇÃO DO SEPTO NASAL - PEDIATRICO": nasal,
                "CORREÇÃO CIRÚRGICA DE ESTRABISMO (ACIMA DE 2 MUSCULOS) - PEDIATRICO": estrabismo,
                "HERNIOPLASTIA INGUINAL (BILATERAL) - PEDIATRICO": inguinal,
                "HERNIOPLASTIA UMBILICAL - PEDIATRICO": umbilical,
                "ORQUIDOPEXIA BILATERAL - PEDIATRICO": orqui,
                "TRATAMENTO CIRÚRGICO DE HIDROCELE - PEDIATRICO": hidrocele,
                "CORRECAO DE HIPOSPADIA (1º TEMPO) - PEDIATRICO": hispospadia,
                "PLASTICA TOTAL DO PENIS - PEDIATRICO": plastica,
                "POSTECTOMIA - PEDIATRICO": postec,
            }
            
            totais = {}
            for nome, lista in grupos.items():
                totais[nome] = process_group(lista or [])
        
            resultados_por_municipio[municipio] = totais
            
            print(f"=== RESULTADOS {municipio} ===")
            for nome, total in totais.items():
                print(f"{nome}: {total}")
            soma_total = sum(totais.values())
            print(f"Soma Total: {soma_total}")
            print("\n")
            
            # Procedimentos não listados
            todos_procedimentos_conhecidos = []
            for lista in grupos.values():
                todos_procedimentos_conhecidos.extend(lista or [])
            
            todos_procedimentos_conhecidos = list(set([str(p) for p in todos_procedimentos_conhecidos if p and str(p).strip()]))
            
            procedimentos_na_coluna = tabela[coluna_procedimento].dropna().unique()
            procedimentos_na_coluna = [str(p) for p in procedimentos_na_coluna if p and str(p).strip()]
            
            procedimentos_nao_mapeados = {}
            for procedimento in procedimentos_na_coluna:
                if procedimento not in todos_procedimentos_conhecidos:
                    mask = tabela[coluna_procedimento].astype(str) == str(procedimento)
                    quantidade = tabela.loc[mask, coluna_quantidade].sum()
                    
                    try:
                        quantidade_num = float(quantidade) if not pd.isna(quantidade) else 0
                    except (ValueError, TypeError):
                        quantidade_num = 0
                    
                    if quantidade_num > 0:
                        procedimentos_nao_mapeados[procedimento] = quantidade_num

            if procedimentos_nao_mapeados:
                procedimentos_nao_mapeados = dict(sorted(procedimentos_nao_mapeados.items(), 
                                                        key=lambda x: x[1], reverse=True))
                
                print(f"Encontrados {len(procedimentos_nao_mapeados)} procedimentos não mapeados em {municipio}")
                
                # Criar DataFrame para não listados deste município
                df_nao_listado = pd.DataFrame({
                    "Procedimento": list(procedimentos_nao_mapeados.keys()),
                    "Quantidade": list(procedimentos_nao_mapeados.values()),
                    "Município": municipio
                })
                nao_listados_dfs.append(df_nao_listado)
            
        except FileNotFoundError:
            print(f"Arquivo {arquivo} não encontrado!")
            continue  

        except Exception as e:
            print(f"Erro em {municipio}: {e}")
            import traceback
            traceback.print_exc()
            continue

    # Salvar resultados consolidados
    if resultados_por_municipio: 
        todos_procedimentos = set()
        for municipio, totais in resultados_por_municipio.items():
            todos_procedimentos.update(totais.keys())
        
        todos_procedimentos = sorted(list(todos_procedimentos))
        
        df_consolidado = pd.DataFrame(index=todos_procedimentos)
        
        # Adicionar colunas para cada município
        for municipio in municipios:
            totais = resultados_por_municipio.get(municipio, {})

            valores = []
            for procedimento in todos_procedimentos:
                valores.append(totais.get(procedimento, 0))
            
            df_consolidado[municipio] = valores
        
        # Adicionar linha de totais
        df_consolidado.loc['TOTAL'] = df_consolidado.sum()
        
        # Salvar em Excel
        output_path = output_dir / "TODOS_MUNICIPIOS_CONSOLIDADO.xlsx"
        df_consolidado.to_excel(output_path)
        print(f"\nArquivo consolidado salvo em: {output_path}")
        print(f"Dimensões: {df_consolidado.shape}")
    
    # CRIAR EXCEL COM PROCEDIMENTOS NÃO LISTADOS
    if nao_listados_dfs:
        # Juntar todos os DataFrames verticalmente
        df_nao_listados_consolidado = pd.concat(nao_listados_dfs, ignore_index=True)
        
        # Reorganizar colunas
        df_nao_listados_consolidado = df_nao_listados_consolidado[["Município", "Procedimento", "Quantidade"]]
        
        output_nao_listados_path = output_dir / "PROCEDIMENTOS_NAO_LISTADOS_CONSOLIDADO.xlsx"
        df_nao_listados_consolidado.to_excel(output_nao_listados_path, index=False)
        print(f"Arquivo de procedimentos não listados salvo em: {output_nao_listados_path}")
        print(f"Total de procedimentos não listados: {len(df_nao_listados_consolidado)}")

# ============================================================================
# FUNÇÃO MAIN PRINCIPAL
# ============================================================================

def main():
    """
    Função principal que coordena todo o processamento
    """
    print("=" * 80)
    print("INICIANDO ANÁLISE DOS RELATÓRIOS")
    print("=" * 80)
    
    # Primeiro, processar os arquivos para criar as versões simplificadas
    print("\n1. Processando arquivos originais...")
    processar_todos_arquivos_simplificado()
    
    # Depois, analisar o Neomater
    print("\n2. Analisando dados do Neomater...")
    analisar_neomater()
    
    print("\n" + "=" * 80)
    print("PROCESSAMENTO CONCLUÍDO!")
    print("=" * 80)

# ============================================================================
# EXECUÇÃO PRINCIPAL
# ============================================================================

if __name__ == "__main__":
    # Este código será executado quando rodar o script diretamente
    # Mas quando usado como módulo pela interface gráfica, só chamará as funções
    main()